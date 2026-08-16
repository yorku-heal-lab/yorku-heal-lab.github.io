#!/usr/bin/env python3
"""Sync resources from the resources spreadsheet into site data."""

from __future__ import annotations

import argparse
import re
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "scripts" / "data" / "resources.xlsx"
DEFAULT_RESOURCES_YML = ROOT / "_data" / "resources.yml"

COLUMNS = {
    "section": "section",
    "name": "name",
    "description": "description",
    "url": "url",
    "tags": "tags",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Input resources spreadsheet path")
    parser.add_argument(
        "--resources-yml",
        type=Path,
        default=DEFAULT_RESOURCES_YML,
        help="Output resources data file",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and summarize without writing resources.yml",
    )
    return parser.parse_args()


def setup_yaml() -> None:
    def represent_str(dumper: yaml.Dumper, data: str) -> yaml.nodes.ScalarNode:
        if "\n" in data:
            return dumper.represent_scalar("tag:yaml.org,2002:str", data, style=">")
        if len(data) > 100:
            return dumper.represent_scalar("tag:yaml.org,2002:str", data, style="|")
        return dumper.represent_scalar("tag:yaml.org,2002:str", data)

    yaml.add_representer(str, represent_str)


def dump_yaml(path: Path, data: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        yaml.dump(data, handle, sort_keys=False, allow_unicode=True, width=1000)


def normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_tags(value: object) -> list[str]:
    text = normalize(value)
    if not text:
        return []
    chunks = re.split(r"[;\n,]+", text)
    return [chunk.strip() for chunk in chunks if chunk.strip()]


def read_resource_rows(xlsx_path: Path) -> tuple[list[dict], list[str]]:
    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook.active

    headers = [normalize(cell.value).lower() for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    header_index = {header: idx for idx, header in enumerate(headers)}

    missing = [label for label in COLUMNS.values() if label not in header_index]
    if missing:
        raise ValueError(f"Missing expected columns in spreadsheet: {', '.join(missing)}")

    items: list[dict] = []
    warnings: list[str] = []

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        record = {
            key: row[header_index[label]] if header_index[label] < len(row) else None
            for key, label in COLUMNS.items()
        }

        section = normalize(record["section"])
        name = normalize(record["name"])
        description = normalize(record["description"])
        url = normalize(record["url"])
        tags = parse_tags(record["tags"])

        if not section and not name and not description and not url:
            continue

        if not section or not name or not description or not url:
            warnings.append(f"Row {row_number}: skipped entry missing section, name, description, or url")
            continue

        item = {
            "section": section,
            "name": name,
            "description": description,
            "url": url,
        }
        if tags:
            item["tags"] = tags

        items.append(item)

    return items, warnings


def build_sections(items: list[dict]) -> list[dict]:
    grouped: OrderedDict[str, list[dict]] = OrderedDict()

    for item in items:
        section_title = item["section"]
        resource = {
            "name": item["name"],
            "description": item["description"],
            "url": item["url"],
        }
        if item.get("tags"):
            resource["tags"] = item["tags"]

        grouped.setdefault(section_title, []).append(resource)

    sections = []
    for title, section_items in grouped.items():
        section_items.sort(key=lambda resource: resource["name"].casefold())
        sections.append({"title": title, "items": section_items})

    return sections


def main() -> int:
    setup_yaml()
    args = parse_args()

    if not args.xlsx.exists():
        print(f"Spreadsheet not found: {args.xlsx}", file=sys.stderr)
        return 1

    items, warnings = read_resource_rows(args.xlsx)
    sections = build_sections(items)

    print(f"Parsed {len(items)} resource(s) in {len(sections)} section(s) from {args.xlsx.name}")

    if warnings:
        print(f"Warnings ({len(warnings)}):", file=sys.stderr)
        for warning in warnings:
            print(f"  - {warning}", file=sys.stderr)

    if not sections:
        print("No resources parsed. Existing resources.yml was not changed.", file=sys.stderr)
        return 1

    if args.dry_run:
        print("Dry run complete; no files written.")
        for section in sections:
            print(f"  [{section['title']}] {len(section['items'])} item(s)")
        return 0

    dump_yaml(args.resources_yml, {"sections": sections})
    print(f"Updated {args.resources_yml.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
