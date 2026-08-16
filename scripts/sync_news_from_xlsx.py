#!/usr/bin/env python3
"""Sync news items from the news spreadsheet into site data."""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "scripts" / "data" / "news.xlsx"
DEFAULT_NEWS_YML = ROOT / "_data" / "news.yml"

COLUMNS = {
    "date": "date",
    "title": "title",
    "summary": "summary",
    "link": "link",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Input news spreadsheet path")
    parser.add_argument("--news-yml", type=Path, default=DEFAULT_NEWS_YML, help="Output news data file")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and summarize without writing news.yml",
    )
    return parser.parse_args()


def setup_yaml() -> None:
    def represent_str(dumper: yaml.Dumper, data: str) -> yaml.nodes.ScalarNode:
        if "\n" in data:
            return dumper.represent_scalar("tag:yaml.org,2002:str", data, style=">")
        if len(data) > 100:
            return dumper.represent_scalar("tag:yaml.org,2002:str", data, style="|")
        return dumper.represent_scalar("tag:yaml.org,2002:str", data)

    yaml.add_representer(str, represent_str)


def dump_yaml(path: Path, data: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        yaml.dump(data, handle, sort_keys=False, allow_unicode=True, width=1000)


def normalize(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def parse_date(value: object) -> str:
    text = normalize(value)
    if not text:
        return ""

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue

    match = re.search(r"(\d{4})-(\d{2})-(\d{2})", text)
    if match:
        return match.group(0)

    raise ValueError(f"Could not parse date: {text!r}")


def read_news_rows(xlsx_path: Path) -> tuple[list[dict], list[str]]:
    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook.active

    headers = [normalize(cell.value).lower() for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    header_index = {header: idx for idx, header in enumerate(headers)}

    missing = [label for label in COLUMNS.values() if label not in header_index]
    if missing:
        raise ValueError(f"Missing expected columns in spreadsheet: {', '.join(missing)}")

    items: list[dict] = []
    warnings: list[str] = []

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        record = {
            key: row[header_index[label]] if header_index[label] < len(row) else None
            for key, label in COLUMNS.items()
        }

        title = normalize(record["title"])
        summary = normalize(record["summary"])
        link = normalize(record["link"])

        if not title and not summary:
            continue

        if not title or not summary:
            warnings.append(f"Row {row_number}: skipped entry missing title or summary")
            continue

        try:
            parsed_date = parse_date(record["date"])
        except ValueError as exc:
            warnings.append(f"Row {row_number}: {exc}")
            continue

        item = {
            "date": parsed_date,
            "title": title,
            "summary": summary,
        }
        if link:
            item["link"] = link

        items.append(item)

    return items, warnings


def sort_news(items: list[dict]) -> list[dict]:
    return sorted(items, key=lambda item: item["date"], reverse=True)


def main() -> int:
    setup_yaml()
    args = parse_args()

    if not args.xlsx.exists():
        print(f"Spreadsheet not found: {args.xlsx}", file=sys.stderr)
        return 1

    items, warnings = read_news_rows(args.xlsx)
    items = sort_news(items)

    print(f"Parsed {len(items)} news item(s) from {args.xlsx.name}")

    if warnings:
        print(f"Warnings ({len(warnings)}):", file=sys.stderr)
        for warning in warnings:
            print(f"  - {warning}", file=sys.stderr)

    if not items:
        print("No news items parsed. Existing news.yml was not changed.", file=sys.stderr)
        return 1

    if args.dry_run:
        print("Dry run complete; no files written.")
        for item in items[:5]:
            print(f"  [{item['date']}] {item['title']}")
        if len(items) > 5:
            print(f"  ... and {len(items) - 5} more")
        return 0

    dump_yaml(args.news_yml, items)
    print(f"Updated {args.news_yml.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
