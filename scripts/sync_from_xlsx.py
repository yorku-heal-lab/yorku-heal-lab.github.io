#!/usr/bin/env python3
"""Sync co-directors and team members from the HEAL lab spreadsheet into site data."""

from __future__ import annotations

import argparse
import re
import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "scripts" / "data" / "HEAL lab info for website.xlsx"
DEFAULT_LAB_YML = ROOT / "_data" / "lab.yml"
DEFAULT_TEAM_YML = ROOT / "_data" / "team.yml"
TEAM_IMAGES_DIR = ROOT / "assets" / "images" / "team"
PLACEHOLDER_PHOTO = "/assets/images/team/placeholder.svg"

TEAM_GROUPS = ("postdocs", "phd_students", "masters_students", "collaborators")

COLUMNS = {
    "name": "name to display",
    "role_in_lab": "role in the lab",
    "role_outside": "role outside (if any)",
    "bio": "short bio",
    "publications": "publication list (at Lab)",
    "email": "York contact info (email)",
    "interests": "research interests (keywords)",
    "image_hint": "link to profile image to use",
}

DEFAULT_DEPARTMENT = "School of Health Policy and Management"
DEFAULT_FACULTY = "Faculty of Health, York University"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Input spreadsheet path")
    parser.add_argument("--lab-yml", type=Path, default=DEFAULT_LAB_YML, help="Output lab data file")
    parser.add_argument("--team-yml", type=Path, default=DEFAULT_TEAM_YML, help="Output team data file")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print summary without writing files or copying images",
    )
    return parser.parse_args()


def setup_yaml() -> None:
    def represent_str(dumper: yaml.Dumper, data: str) -> yaml.nodes.ScalarNode:
        if "\n" in data:
            return dumper.represent_scalar("tag:yaml.org,2002:str", data, style=">")
        if len(data) > 100:
            return dumper.represent_scalar("tag:yaml.org,2002:str", data, style="|")
        return dumper.represent_scalar("tag:yaml.org,2002:str", data)

    yaml.add_representer(str, represent_str)


def load_yaml(path: Path) -> dict:
    if not path.exists():
        return {}
    with path.open(encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def dump_yaml(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        yaml.dump(data, handle, sort_keys=False, allow_unicode=True, width=1000)


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "member"


def normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def first_name(name: str) -> str:
    cleaned = re.sub(r"^Dr\.?\s+", "", name, flags=re.IGNORECASE).strip()
    parts = cleaned.split()
    return parts[0].casefold() if parts else cleaned.casefold()


def sort_by_first_name(members: list[dict]) -> list[dict]:
    return sorted(members, key=lambda member: (first_name(member["name"]), member["name"].casefold()))


def is_co_director(role_in_lab: str) -> bool:
    return "co-director" in role_in_lab.lower()


def classify_team_group(role_in_lab: str) -> str | None:
    lowered = role_in_lab.lower()

    if is_co_director(lowered):
        return None
    if "collaborator" in lowered:
        return "collaborators"
    if "postdoc" in lowered or "postdoctoral" in lowered:
        return "postdocs"
    if re.search(r"\bph\.?\s*d\.?\b", lowered) or "phd" in lowered or "doctoral" in lowered:
        return "phd_students"
    if re.search(r"\bm\.?\s*(sc|a)\.?\b", lowered) or "master" in lowered or "msc" in lowered:
        return "masters_students"

    return None


def display_name(name: str, bio: str, role_in_lab: str) -> str:
    if name.startswith("Dr."):
        return name
    if bio.startswith("Dr.") and name.split()[0] in bio.split()[0:3]:
        return f"Dr. {name}"
    lowered = role_in_lab.lower()
    if any(token in lowered for token in ("professor", "postdoctoral", "director")):
        return f"Dr. {name}"
    return name


def parse_director_title(role_in_lab: str) -> str:
    parts = [part.strip() for part in role_in_lab.split(",") if part.strip()]
    if not parts:
        return role_in_lab
    if is_co_director(parts[0]):
        remaining = parts[1:]
        return ", ".join(remaining) if remaining else parts[0]
    return role_in_lab


def first_paragraph(text: str) -> str:
    paragraph = text.split("\n\n", 1)[0].strip()
    return re.sub(r"\s+", " ", paragraph)


def short_bio(text: str, limit: int = 320) -> str:
    paragraph = first_paragraph(text)
    if len(paragraph) <= limit:
        return paragraph
    trimmed = paragraph[: limit - 1].rsplit(" ", 1)[0]
    return f"{trimmed}..."


def parse_interests(text: str) -> list[str]:
    if not text:
        return []
    chunks = re.split(r"[;\n]+", text)
    return [chunk.strip() for chunk in chunks if chunk.strip()]


def read_rows(xlsx_path: Path) -> tuple[list[str], list[dict[str, str]], dict[int, str]]:
    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook.active

    headers = [normalize(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    header_index = {header: idx for idx, header in enumerate(headers)}

    missing = [label for label in COLUMNS.values() if label not in header_index]
    if missing:
        raise ValueError(f"Missing expected columns in spreadsheet: {', '.join(missing)}")

    rows: list[dict[str, str]] = []
    for excel_row, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        record = {
            key: normalize(row[header_index[label]]) for key, label in COLUMNS.items()
        }
        record["excel_row"] = str(excel_row)
        if record["name"]:
            rows.append(record)

    row_media = build_row_media_map(xlsx_path)
    return headers, rows, row_media


def media_extension(path: str) -> str:
    suffix = Path(path).suffix.lower()
    return suffix if suffix in {".jpg", ".jpeg", ".png", ".svg", ".webp", ".gif"} else ".jpg"


def build_row_media_map(xlsx_path: Path) -> dict[int, str]:
    with zipfile.ZipFile(xlsx_path) as archive:
        drawing = archive.read("xl/drawings/drawing1.xml").decode("utf-8")
        rels = archive.read("xl/drawings/_rels/drawing1.xml.rels").decode("utf-8")

        rel_map = {
            match.group(1): match.group(2)
            for match in re.finditer(r'Id="([^"]+)".*?Target="([^"]+)"', rels)
        }

        row_media: dict[int, str] = {}
        anchors = re.findall(
            r"<xdr:from>.*?<xdr:row>(\d+)</xdr:row>.*?r:embed=\"([^\"]+)\"",
            drawing,
            flags=re.S,
        )
        for row_index, rel_id in anchors:
            target = rel_map.get(rel_id, "")
            if target:
                excel_row = int(row_index) + 1
                row_media[excel_row] = f"xl/{target.removeprefix('../')}"

        return row_media


def extract_image_bytes(xlsx_path: Path, media_path: str) -> bytes:
    with zipfile.ZipFile(xlsx_path) as archive:
        if media_path in archive.namelist():
            return archive.read(media_path)
        basename = Path(media_path).name
        for member in archive.namelist():
            if member.endswith("/" + basename):
                return archive.read(member)
        raise FileNotFoundError(f"Could not find image media '{media_path}' in {xlsx_path}")


def resolve_photo_path(
    record: dict[str, str],
    row_media: dict[int, str],
    xlsx_path: Path,
    dry_run: bool,
) -> str:
    excel_row = int(record["excel_row"])
    slug = slugify(record["name"])
    media_path = row_media.get(excel_row)

    if media_path is not None:
        extension = media_extension(media_path)
        destination = TEAM_IMAGES_DIR / f"{slug}{extension}"
        web_path = f"/assets/images/team/{destination.name}"
        if dry_run:
            return web_path
        TEAM_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
        destination.write_bytes(extract_image_bytes(xlsx_path, media_path))
        return web_path

    image_hint = record["image_hint"]
    if image_hint:
        hint_path = xlsx_path.parent / image_hint
        if hint_path.exists():
            destination = TEAM_IMAGES_DIR / f"{slug}{hint_path.suffix.lower()}"
            web_path = f"/assets/images/team/{destination.name}"
            if dry_run:
                return web_path
            TEAM_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
            shutil.copy2(hint_path, destination)
            return web_path

    return PLACEHOLDER_PHOTO


def build_director(record: dict[str, str], photo: str) -> dict:
    role_in_lab = record["role_in_lab"]
    bio = record["bio"]
    director = {
        "name": display_name(record["name"], bio, role_in_lab),
        "role": "Co-Director",
        "title": parse_director_title(role_in_lab),
        "department": DEFAULT_DEPARTMENT,
        "faculty": DEFAULT_FACULTY,
        "photo": photo,
    }
    if bio:
        director["short_bio"] = short_bio(bio)
    if record["email"]:
        director["email"] = record["email"]
    if record["publications"]:
        director["website"] = record["publications"]
    return director


def build_team_member(record: dict[str, str], photo: str) -> dict:
    role_in_lab = record["role_in_lab"]
    bio = record["bio"]
    member = {
        "name": display_name(record["name"], bio, role_in_lab),
        "role": role_in_lab,
        "photo": photo,
    }
    if record["email"]:
        member["email"] = record["email"]
    if record["role_outside"]:
        member["affiliation"] = record["role_outside"]
    if bio:
        member["bio"] = bio
    return member


def merge_lab_data(existing: dict, directors: list[dict]) -> dict:
    lab_data = existing.copy()
    lab_data["directors"] = directors

    if directors:
        first_email = directors[0].get("email")
        contact = dict(lab_data.get("contact") or {})
        if first_email:
            contact["email"] = first_email
        lab_data["contact"] = contact

    return lab_data


def merge_team_data(existing: dict, grouped_members: dict[str, list[dict]]) -> dict:
    team_data = {
        group: sort_by_first_name(grouped_members.get(group, [])) for group in TEAM_GROUPS
    }
    team_data["alumni"] = sort_by_first_name(existing.get("alumni") or [])
    return team_data


def empty_team_groups() -> dict[str, list[dict]]:
    return {group: [] for group in TEAM_GROUPS}


def main() -> int:
    setup_yaml()
    args = parse_args()

    if not args.xlsx.exists():
        print(f"Spreadsheet not found: {args.xlsx}", file=sys.stderr)
        return 1

    _, rows, row_media = read_rows(args.xlsx)

    co_directors: list[dict] = []
    team_groups = empty_team_groups()
    uncategorized: list[str] = []

    for record in rows:
        photo = resolve_photo_path(record, row_media, args.xlsx, args.dry_run)
        role_in_lab = record["role_in_lab"]

        if is_co_director(role_in_lab):
            co_directors.append(build_director(record, photo))
            continue

        group = classify_team_group(role_in_lab)
        if group is None:
            uncategorized.append(f"{record['name']} ({role_in_lab})")
            continue

        team_groups[group].append(build_team_member(record, photo))

    lab_data = merge_lab_data(load_yaml(args.lab_yml), sort_by_first_name(co_directors))
    team_data = merge_team_data(load_yaml(args.team_yml), team_groups)

    print(f"Parsed {len(rows)} people from {args.xlsx.name}")
    print(f"Co-directors ({len(co_directors)}): {', '.join(d['name'] for d in co_directors) or 'none'}")
    for group in TEAM_GROUPS:
        members = team_groups[group]
        label = group.replace("_", " ")
        print(f"{label.title()} ({len(members)}): {', '.join(m['name'] for m in members) or 'none'}")
    if uncategorized:
        print("Uncategorized (skipped): " + "; ".join(uncategorized), file=sys.stderr)

    if args.dry_run:
        print("Dry run complete; no files written.")
        return 0

    dump_yaml(args.lab_yml, lab_data)
    dump_yaml(args.team_yml, team_data)
    print(f"Updated {args.lab_yml.relative_to(ROOT)}")
    print(f"Updated {args.team_yml.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
