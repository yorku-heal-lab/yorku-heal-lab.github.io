#!/usr/bin/env python3
"""Sync research content from the research spreadsheet into site data."""

from __future__ import annotations

import argparse
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "scripts" / "data" / "research.xlsx"
DEFAULT_RESEARCH_YML = ROOT / "_data" / "research.yml"

COLUMNS = {
    "section": "section",
    "format": "format",
    "title": "title",
    "description": "description",
    "url": "url",
}

VALID_FORMATS = {"theme", "list"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Input research spreadsheet path")
    parser.add_argument(
        "--research-yml",
        type=Path,
        default=DEFAULT_RESEARCH_YML,
        help="Output research data file",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and summarize without writing research.yml",
    )
    return parser.parse_args()


def setup_yaml() -> None:
    def represent_str(dumper: yaml.Dumper, data: str) -> yaml.nodes.ScalarNode:
        if "\n" in data:
            return dumper.represent_scalar("tag:yaml.org,2002:str", data, style=">")
        if len(data) > 100:
            return dumper.represent_scalar("tag:yaml.org,2002:str", data, style="|")
        return dumper.represent_scalar("tag:yaml.org,2002:str", data)

    yaml.add_representer(str, represent_str)


def dump_yaml(path: Path, data: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        yaml.dump(data, handle, sort_keys=False, allow_unicode=True, width=1000)


def normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def default_format(section: str) -> str:
    if "project" in section.casefold():
        return "list"
    return "theme"


def normalize_format(value: object, section: str) -> str:
    text = normalize(value).lower()
    if not text:
        return default_format(section)
    if text not in VALID_FORMATS:
        raise ValueError(f"Unsupported format {text!r}; expected one of: {', '.join(sorted(VALID_FORMATS))}")
    return text


def read_research_rows(xlsx_path: Path) -> tuple[list[dict], list[str]]:
    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook.active

    headers = [normalize(cell.value).lower() for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    header_index = {header: idx for idx, header in enumerate(headers)}

    missing = [label for label in COLUMNS.values() if label not in header_index]
    if missing:
        raise ValueError(f"Missing expected columns in spreadsheet: {', '.join(missing)}")

    items: list[dict] = []
    warnings: list[str] = []

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        record = {
            key: row[header_index[label]] if header_index[label] < len(row) else None
            for key, label in COLUMNS.items()
        }

        section = normalize(record["section"])
        title = normalize(record["title"])
        description = normalize(record["description"])
        url = normalize(record["url"])

        if not section and not title and not description:
            continue

        if not section or not title:
            warnings.append(f"Row {row_number}: skipped entry missing section or title")
            continue

        try:
            item_format = normalize_format(record["format"], section)
        except ValueError as exc:
            warnings.append(f"Row {row_number}: {exc}")
            continue

        item = {
            "section": section,
            "format": item_format,
            "title": title,
        }
        if description:
            item["description"] = description
        if url:
            item["url"] = url

        items.append(item)

    return items, warnings


def build_sections(items: list[dict]) -> list[dict]:
    grouped: OrderedDict[str, dict] = OrderedDict()

    for item in items:
        section_title = item["section"]
        if section_title not in grouped:
            grouped[section_title] = {
                "title": section_title,
                "format": item["format"],
                "items": [],
            }

        entry = {"title": item["title"]}
        if item.get("description"):
            entry["description"] = item["description"]
        if item.get("url"):
            entry["url"] = item["url"]
        grouped[section_title]["items"].append(entry)

    sections = []
    for section in grouped.values():
        section["items"].sort(key=lambda entry: entry["title"].casefold())
        sections.append(section)

    return sections


def main() -> int:
    setup_yaml()
    args = parse_args()

    if not args.xlsx.exists():
        print(f"Spreadsheet not found: {args.xlsx}", file=sys.stderr)
        return 1

    items, warnings = read_research_rows(args.xlsx)
    sections = build_sections(items)

    print(f"Parsed {len(items)} research item(s) in {len(sections)} section(s) from {args.xlsx.name}")

    if warnings:
        print(f"Warnings ({len(warnings)}):", file=sys.stderr)
        for warning in warnings:
            print(f"  - {warning}", file=sys.stderr)

    if not sections:
        print("No research content parsed. Existing research.yml was not changed.", file=sys.stderr)
        return 1

    if args.dry_run:
        print("Dry run complete; no files written.")
        for section in sections:
            print(f"  [{section['title']}] ({section['format']}) {len(section['items'])} item(s)")
        return 0

    dump_yaml(args.research_yml, {"sections": sections})
    print(f"Updated {args.research_yml.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
